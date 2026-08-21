#!/usr/bin/env python3
"""Build and visually review the bundled v30 memory-item icon set.

The source directory contains 7x7 transparent PNG montages named by their
Excel row span (for example, 2-50.png). Each connected alpha component is
assigned to exactly one grid cell by its center, so detached details stay with
their icon and pixels from neighboring cells cannot leak into the output.

Outputs:
  apps/mobile/assets/items/each/memory.<icon-name>.webp
  tools/item-source/memory-items/standardized-preview.html
  tools/item-source/memory-items/items-v30-image-qa.json

Normal runs are incremental: only workbook items whose WebP is missing are
decoded and written, so previously bundled artwork is never touched. Pass
--full only when an intentional complete rebuild is required.
"""

from __future__ import annotations

import argparse
import html
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from PIL import Image
from scipy import ndimage


ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "tools" / "item-source" / "memory-items"
WORKBOOK = SOURCE_DIR / "Icon_Mapping_Core_Tables_v30.xlsx"
OUTPUT_DIR = ROOT / "apps" / "mobile" / "assets" / "items" / "each"
PREVIEW_PATH = SOURCE_DIR / "standardized-preview.html"
QA_PATH = SOURCE_DIR / "items-v30-image-qa.json"

GRID = 7
CANVAS = 256
SAFE_SIZE = 200
ALPHA_THRESHOLD = 2
ICON_COUNT = 5439
SOURCE_PAGE_COUNT = 111
FINAL_WORKBOOK_ROW = 5440
SOURCE_RE = re.compile(r"^(\d+)-(\d+)\.png$")
FORCE_PROJECTION_PAGES = {"2256-2304.png", "5343-5391.png"}


def slugify(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    value = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    if not value:
        raise ValueError("Icon_name produced an empty slug")
    return value


def source_pages() -> list[tuple[int, int, Path]]:
    pages = []
    for path in SOURCE_DIR.glob("*.png"):
        match = SOURCE_RE.match(path.name)
        if match:
            start, end = map(int, match.groups())
            pages.append((start, end, path))
    pages.sort(key=lambda page: page[0])
    if len(pages) != SOURCE_PAGE_COUNT:
        raise ValueError(f"Expected {SOURCE_PAGE_COUNT} source pages, found {len(pages)}")
    expected = 2
    for start, end, path in pages:
        if start != expected or end - start + 1 != GRID * GRID:
            raise ValueError(f"Invalid or non-contiguous page range: {path.name}")
        expected = end + 1
    if expected != FINAL_WORKBOOK_ROW + 1:
        raise ValueError(f"Expected final workbook row {FINAL_WORKBOOK_ROW}, got {expected - 1}")
    return pages


def workbook_items() -> dict[int, dict[str, str]]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["Icon_Mapping"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    indexes = {name: index for index, name in enumerate(headers) if name}
    required = {"Icon_name", "Primary_Domain", "Tier", "Status"}
    missing = required - indexes.keys()
    if missing:
        raise ValueError(f"Icon_Mapping missing columns: {sorted(missing)}")

    result: dict[int, dict[str, str]] = {}
    seen_names: set[str] = set()
    seen_ids: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        name = str(row[indexes["Icon_name"]] or "").strip()
        if not name:
            raise ValueError(f"Missing Icon_name at workbook row {row_number}")
        if name.casefold() in seen_names:
            raise ValueError(f"Duplicate Icon_name: {name}")
        item_id = f"memory.{row_number:04d}_{slugify(name)}"
        if item_id in seen_ids:
            raise ValueError(f"Duplicate generated item id: {item_id}")
        seen_names.add(name.casefold())
        seen_ids.add(item_id)
        result[row_number] = {
            "name": name,
            "item_id": item_id,
            "primary_domain": str(row[indexes["Primary_Domain"]] or "Uncategorized").strip(),
            "tier": str(row[indexes["Tier"]] or "").strip(),
            "status": str(row[indexes["Status"]] or "").strip(),
        }
    if len(result) != ICON_COUNT:
        raise ValueError(f"Expected {ICON_COUNT} workbook icons, found {len(result)}")
    return result


def projection_runs(on: np.ndarray) -> list[tuple[int, int]]:
    runs: list[tuple[int, int]] = []
    start = None
    for index, occupied in enumerate(on.tolist() + [False]):
        if occupied and start is None:
            start = index
        elif not occupied and start is not None:
            runs.append((start, index))
            start = None
    return runs


def projection_groups(on: np.ndarray, target: int, label: str) -> list[tuple[int, int]]:
    """Group occupied projection runs by retaining the largest target-1 gaps."""
    runs = projection_runs(on)
    if len(runs) < target:
        raise ValueError(f"{label}: only {len(runs)} occupied runs for {target} groups")
    while len(runs) > target:
        merge_at = min(
            range(len(runs) - 1),
            key=lambda index: runs[index + 1][0] - runs[index][1],
        )
        runs[merge_at : merge_at + 2] = [(runs[merge_at][0], runs[merge_at + 1][1])]
    return runs


def render_icon(source_rgba: np.ndarray, own_mask: np.ndarray) -> tuple[Image.Image, tuple[int, int, int, int]]:
    ys, xs = np.where(own_mask)
    if not len(xs):
        raise ValueError("Cannot render an empty icon mask")
    y0, y1 = int(ys.min()), int(ys.max()) + 1
    x0, x1 = int(xs.min()), int(xs.max()) + 1
    patch = source_rgba[y0:y1, x0:x1].copy()
    patch[~own_mask[y0:y1, x0:x1], 3] = 0
    icon = Image.fromarray(patch, "RGBA")
    scale = min(SAFE_SIZE / icon.width, SAFE_SIZE / icon.height)
    resized = icon.resize(
        (max(1, round(icon.width * scale)), max(1, round(icon.height * scale))),
        Image.Resampling.LANCZOS,
    )
    canvas = Image.new("RGBA", (CANVAS, CANVAS), (0, 0, 0, 0))
    canvas.alpha_composite(
        resized,
        ((CANVAS - resized.width) // 2, (CANVAS - resized.height) // 2),
    )
    bbox = canvas.getchannel("A").getbbox()
    if bbox is None:
        raise ValueError("Generated an empty icon")
    return canvas, bbox


def extract_page(path: Path) -> tuple[list[Image.Image], dict[str, object]]:
    source = Image.open(path).convert("RGBA")
    array = np.asarray(source)
    alpha = array[..., 3]
    raw_content = alpha > ALPHA_THRESHOLD
    raw_labels, raw_component_count = ndimage.label(
        raw_content, structure=np.ones((3, 3), dtype=int)
    )
    component_sizes = np.bincount(raw_labels.ravel())
    kept_components = np.where(component_sizes >= 16)[0]
    kept_components = kept_components[kept_components != 0]
    content = np.isin(raw_labels, kept_components)
    _, component_count = ndimage.label(content, structure=np.ones((3, 3), dtype=int))
    height, width = content.shape
    source_rgba = array.copy()

    # Most pages become exactly 49 icon bodies after zero or a few pixels of
    # dilation. This is the strongest isolation method: each source pixel can
    # belong to only one labeled body, even when the visual rows overlap.
    for dilation in ([] if path.name in FORCE_PROJECTION_PAGES else range(0, 9)):
        grouped = content if dilation == 0 else ndimage.binary_dilation(content, iterations=dilation)
        group_labels, group_count = ndimage.label(
            grouped, structure=np.ones((3, 3), dtype=int)
        )
        if group_count != GRID * GRID:
            continue
        records = []
        for component_id, slices in enumerate(ndimage.find_objects(group_labels), start=1):
            if slices is None:
                continue
            y0, y1 = slices[0].start, slices[0].stop
            x0, x1 = slices[1].start, slices[1].stop
            records.append((component_id, (y0 + y1) / 2, (x0 + x1) / 2))
        records.sort(key=lambda record: record[1])
        ordered = []
        for row in range(GRID):
            row_records = records[row * GRID : (row + 1) * GRID]
            ordered.extend(sorted(row_records, key=lambda record: record[2]))
        icons, alpha_boxes = [], []
        for component_id, _, _ in ordered:
            own_mask = content & (group_labels == component_id)
            icon, bbox = render_icon(source_rgba, own_mask)
            icons.append(icon)
            alpha_boxes.append(bbox)
        return icons, {
            "source": path.name,
            "width": width,
            "height": height,
            "components": component_count,
            "discarded_tiny_components": raw_component_count - component_count,
            "icons": len(icons),
            "isolation_method": "connected-components",
            "dilation": dilation,
            "minimum_source_row_gap": None,
            "minimum_source_column_gap": None,
            "minimum_margin": min(
                min(x0, y0, CANVAS - x1, CANVAS - y1) for x0, y0, x1, y1 in alpha_boxes
            ),
        }

    # Source pages have transparent gaps between the seven visual rows and
    # between the seven icons within each row. Keeping the six largest gaps is
    # adaptive to intentionally wide icons and uneven placement; fixed 1/7
    # boundaries can put a wide icon into its neighbor's nominal cell.
    if len(projection_runs(content.any(axis=1))) >= GRID:
        row_groups = projection_groups(content.any(axis=1), GRID, f"{path.name} rows")
    else:
        row_groups = []
        row_boundaries = [round(index * height / GRID) for index in range(GRID + 1)]
        for row in range(GRID):
            by0, by1 = row_boundaries[row : row + 2]
            ys = np.where(content[by0:by1].any(axis=1))[0]
            if not len(ys):
                raise ValueError(f"{path.name}: fallback row {row + 1} is empty")
            row_groups.append((by0 + int(ys.min()), by0 + int(ys.max()) + 1))

    icons: list[Image.Image] = []
    alpha_boxes = []
    column_candidates: list[list[tuple[int, int]] | None] = []
    for row, (y0, y1) in enumerate(row_groups):
        on = content[y0:y1].any(axis=0)
        if len(projection_runs(on)) >= GRID:
            groups = projection_groups(on, GRID, f"{path.name} row {row + 1} columns")
            column_candidates.append(groups)
        else:
            column_candidates.append(None)
    # A few source icons touch their immediate neighbor and therefore appear
    # as one projection run. Those rows use the source sheet's seven equal
    # slots only as a fallback; normally the transparent gaps define tighter
    # adaptive groups.
    fallback_boundaries = [round(index * width / GRID) for index in range(GRID + 1)]

    row_gap_sizes = []
    column_gap_sizes = []
    for row, (y0, y1) in enumerate(row_groups):
        if row:
            row_gap_sizes.append(y0 - row_groups[row - 1][1])
        column_groups = column_candidates[row]
        if column_groups is None:
            column_groups = []
            for column in range(GRID):
                bx0, bx1 = fallback_boundaries[column : column + 2]
                xs = np.where(content[y0:y1, bx0:bx1].any(axis=0))[0]
                if not len(xs):
                    raise ValueError(
                        f"{path.name}: fallback column {column + 1} in row {row + 1} is empty"
                    )
                column_groups.append((bx0 + int(xs.min()), bx0 + int(xs.max()) + 1))
        for column, (x0, x1) in enumerate(column_groups):
            if column:
                column_gap_sizes.append(x0 - column_groups[column - 1][1])
            own_mask = np.zeros_like(content)
            own_mask[y0:y1, x0:x1] = content[y0:y1, x0:x1]
            canvas, bbox = render_icon(source_rgba, own_mask)
            alpha_boxes.append(bbox)
            icons.append(canvas)

    return icons, {
        "source": path.name,
        "width": width,
        "height": height,
        "components": component_count,
        "discarded_tiny_components": raw_component_count - component_count,
        "icons": len(icons),
        "isolation_method": "projection-fallback",
        "dilation": None,
        "minimum_source_row_gap": min(row_gap_sizes),
        "minimum_source_column_gap": min(column_gap_sizes),
        "minimum_margin": min(
            min(x0, y0, CANVAS - x1, CANVAS - y1) for x0, y0, x1, y1 in alpha_boxes
        ),
    }


def preview_html(cards_by_page: list[tuple[str, list[dict[str, str]]]]) -> str:
    sections = []
    for page_name, cards in cards_by_page:
        card_markup = "\n".join(
            "<div class='c' data-name='{query}' data-row='{row}' title='Excel row {row} · {item_id}'>"
            "<img loading='lazy' src='../items/each/{filename}' alt='{name}'>"
            "<span>{name}</span><small>Row {row}</small></div>".format(
                query=html.escape(card["name"].casefold(), quote=True),
                row=card["row"],
                item_id=html.escape(card["item_id"], quote=True),
                filename=html.escape(card["filename"], quote=True),
                name=html.escape(card["name"]),
            )
            for card in cards
        )
        sections.append(
            f"<section><h2>{html.escape(page_name)}</h2><div class='grid'>{card_markup}</div></section>"
        )

    return """<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Memory Items v30 · 5,439 Icons</title>
<style>
*{box-sizing:border-box}body{margin:0;background:#f5f0e8;color:#50351e;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif}
header{position:sticky;top:0;z-index:2;padding:16px 24px;background:rgba(245,240,232,.96);border-bottom:1px solid #d8c9b6;backdrop-filter:blur(8px)}
h1{font-size:22px;margin:0 0 10px}header p{margin:0 0 10px;font-size:13px}input{width:min(460px,100%);padding:10px 13px;border:1px solid #bca98f;border-radius:10px;background:white;font-size:15px}
main{padding:10px 24px 40px}section{scroll-margin-top:130px}h2{margin:26px 0 12px;font-size:18px;color:#6b4526}.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(104px,1fr));gap:10px}
.c{min-height:126px;padding:8px 6px 7px;background:white;border:1px solid #eadfce;border-radius:12px;display:flex;flex-direction:column;align-items:center;justify-content:flex-start;text-align:center;box-shadow:0 1px 2px rgba(77,49,25,.06)}
.c img{width:80px;height:80px;object-fit:contain}.c span{font-size:11px;line-height:1.15;min-height:25px;display:flex;align-items:center;justify-content:center}.c small{font-size:9px;color:#a18c73}.hidden{display:none}
</style></head><body>
<header><h1>Memory Items v30</h1><p id="count">5,439 icons · 111 source sheets · 7×7 each</p><input id="search" type="search" placeholder="Search Icon_name or Excel row…"></header>
<main>""" + "\n".join(sections) + """</main>
<script>
const cards=[...document.querySelectorAll('.c')], count=document.querySelector('#count');
document.querySelector('#search').addEventListener('input',e=>{const q=e.target.value.trim().toLowerCase();let visible=0;cards.forEach(c=>{const show=!q||c.dataset.name.includes(q)||c.dataset.row.includes(q);c.classList.toggle('hidden',!show);if(show)visible++});document.querySelectorAll('section').forEach(s=>s.hidden=![...s.querySelectorAll('.c')].some(c=>!c.classList.contains('hidden'));count.textContent=`${visible.toLocaleString()} of 5,439 icons shown`;});
</script></body></html>"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check-only", action="store_true", help="Validate inputs without writing outputs")
    parser.add_argument(
        "--full",
        action="store_true",
        help="Explicitly rebuild every bundled icon; the default only writes missing outputs",
    )
    args = parser.parse_args()

    pages = source_pages()
    items = workbook_items()
    if args.check_only:
        for page_index, (_, _, path) in enumerate(pages, start=1):
            icons, _ = extract_page(path)
            if len(icons) != GRID * GRID:
                raise ValueError(f"{path.name}: expected 49 icons, found {len(icons)}")
            print(f"[{page_index:03d}/{len(pages)}] validated {path.name}")
        print(f"validated {len(pages)} pages and {len(items)} workbook icons")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    expected_all_files = {f"{item['item_id']}.webp" for item in items.values()}
    missing_rows = {
        row_number
        for row_number, item in items.items()
        if not (OUTPUT_DIR / f"{item['item_id']}.webp").exists()
    }
    if args.full:
        # Destructive behavior is opt-in only. Normal runs never remove or
        # overwrite an already-bundled icon.
        for existing in OUTPUT_DIR.glob("memory.*.webp"):
            existing.unlink()
        rows_to_write = set(items)
        selected_pages = pages
        mode = "full"
    else:
        rows_to_write = missing_rows
        selected_pages = [
            page for page in pages
            if any(row_number in rows_to_write for row_number in range(page[0], page[1] + 1))
        ]
        mode = "incremental"

    page_reports = []
    written_files: set[str] = set()
    for page_index, (start, end, path) in enumerate(selected_pages, start=1):
        icons, report = extract_page(path)
        for offset, icon in enumerate(icons):
            row_number = start + offset
            if row_number not in rows_to_write:
                continue
            item = items[row_number]
            filename = f"{item['item_id']}.webp"
            if filename in written_files:
                raise ValueError(f"Duplicate output filename: {filename}")
            icon.save(OUTPUT_DIR / filename, "WEBP", quality=92, method=4)
            written_files.add(filename)
        report["workbook_rows"] = f"{start}-{end}"
        page_reports.append(report)
        print(
            f"[{page_index:03d}/{len(selected_pages)}] {path.name}: "
            f"{sum(row_number in rows_to_write for row_number in range(start, end + 1))} new icons"
        )

    if len(written_files) != len(rows_to_write):
        raise ValueError(f"Expected {len(rows_to_write)} written files, found {len(written_files)}")
    disk_files = {path.name for path in OUTPUT_DIR.glob("memory.*.webp")}
    unexpected = sorted(disk_files - expected_all_files)
    missing = sorted(expected_all_files - disk_files)
    if unexpected or missing:
        raise ValueError(f"Output directory mismatch: unexpected={unexpected[:5]}, missing={missing[:5]}")

    # Preview metadata is cheap to regenerate and references the existing files;
    # incremental image builds do not need to decode or rewrite old artwork.
    cards_by_page = []
    for start, end, path in pages:
        cards = []
        for row_number in range(start, end + 1):
            item = items[row_number]
            cards.append({
                **item,
                "row": str(row_number),
                "filename": f"{item['item_id']}.webp",
            })
        cards_by_page.append((path.stem, cards))
    PREVIEW_PATH.write_text(preview_html(cards_by_page), encoding="utf-8")
    qa = {
        "schema": "memory-items-v30-image-qa@1",
        "mode": mode,
        "source_pages": len(pages),
        "processed_pages": len(selected_pages),
        "icons_expected": ICON_COUNT,
        "icons_generated": len(written_files),
        "catalog_icons_present": len(disk_files),
        "canvas": f"{CANVAS}x{CANVAS}",
        "safe_content_size": SAFE_SIZE,
        "encoding": "WebP quality=92 method=4 with alpha",
        "minimum_output_margin": (
            min(report["minimum_margin"] for report in page_reports) if page_reports else None
        ),
        "unexpected_output_files": unexpected,
        "missing_output_files": missing,
        "pages": page_reports,
    }
    QA_PATH.write_text(json.dumps(qa, indent=2), encoding="utf-8")
    print(f"{mode} build wrote {len(written_files)} icons; {len(disk_files)} catalog icons present")
    print(f"preview: {PREVIEW_PATH}")
    print(f"qa: {QA_PATH}")


if __name__ == "__main__":
    main()

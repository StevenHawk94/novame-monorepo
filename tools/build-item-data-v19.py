#!/usr/bin/env python3
"""Generate runtime item data from the v32 rules over stable memory icon IDs."""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "tools" / "item-source" / "memory-items"
WORKBOOK = SOURCE_DIR / "Icon_Mapping_Core_Tables_v32.xlsx"
DICTIONARY = ROOT / "packages" / "engine" / "src" / "items" / "dictionary.json"
IMAGE_MAP = ROOT / "apps" / "mobile" / "src" / "lib" / "item-images.g.ts"
GUIDED_CATALOG = ROOT / "apps" / "mobile" / "src" / "lib" / "guided-catalog.g.ts"
QA_PATH = SOURCE_DIR / "items-v32-data-qa.json"
ICON_COUNT = 5439
# v32 changes matching and safety rules without adding or renaming icons.
# The deployed catalog is therefore unchanged; this generator refreshes only
# bundled runtime data and generated TypeScript catalogs.
# Keep the existing artwork map byte-for-byte unchanged on this data-only update.
WRITE_IMAGE_MAP = False


CATEGORY_KEYS = {
    "Emotions & Feelings": "emotion_and_feeling",
    "Food & Drink": "food_drink",
    "Chores & Home Care": "chores_home_care",
    "Health & Self-Care": "self_care_hygiene",
    "Work & Productivity": "work_productivity",
    "Entertainment & Leisure": "entertainment_leisure",
    "Exercise & Movement": "exercise_movement",
    "Social & Relationships": "social_relationships",
    "Travel & Getting Around": "travel_commute",
    "Nature & Outdoors": "nature_outdoors",
    "Learning & Hobbies": "learning_hobbies",
    "Shopping & Errands": "shopping_errands",
}

# Cross-icon executable collisions must resolve to one deterministic, concrete
# carrier. Keep this explicit so future workbook changes fail generation until
# each ambiguity has been product-reviewed.
CONFLICT_WINNERS = {
    "ran on the treadmill": "Treadmill",
    "steel toe boots": "Steel-Toe Boot",
    "suitcases": "Suitcase",
    # v32 adds explicit phrases to broad icons that already have more specific
    # owners. Follow the workbook's specific-carrier suppression notes.
    "coral reef": "Coral Reef",
    "counter stool": "Bar Stool",
    "cruise port": "Cruise Terminal",
    "hiking compass": "Pocket Compass",
    "large rock": "Boulder",
    "pocket compass": "Pocket Compass",
    "rock formation": "Rock Formation",
    "storm cloud": "Rain Cloud",
    "toast with jam": "Jam Toast",
    # Looking at stars does not establish that a telescope was used.
    "looked at the stars": "Star",
    # These specific siblings still have non-executable prose exclusions.
    # Use v32's explicit AUTO carrier; do not bypass those safety restrictions.
    "conch shell": "Shell",
    "rain cloud": "Cloud",
}

# Four Keyword_Safety labels refer to retired names that are not present in
# Icon_Mapping v19. Route them to the closest concrete carrier that does exist.
ORPHAN_ICON_REMAP = {
    "Online Shopping": "Shopping",
    "Rideshare": "Car",
    "Slight Smile": "Neutral",
    "Streaming": "Video",
}

INSTRUCTIONAL_EXCLUSION = re.compile(
    r"\b(do not|don't|exclude|unless|require|only when|context|infer|meaning|usage|reference|refers?)\b",
    re.IGNORECASE,
)


def slugify(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def normalize_phrase(value: object) -> str:
    text = str(value or "").strip().lower().replace("’", "'").replace("–", "-")
    return " ".join(re.findall(r"[a-z0-9']+", text))


def parse_literal_exclusions(value: object) -> tuple[list[str], str | None]:
    raw = str(value or "").strip()
    if not raw or raw == "17316":
        return [], None
    lowered = raw.lower()
    if lowered in {
        "no exclusion needed; direct health-object term.",
        "direct sacred-book term is safe.",
    }:
        return [], None
    if lowered == "direct sacred-book term is safe; exclude figurative 'bible of' only when clearly not an object mention.":
        return ["bible of"], None
    parts = [part.strip() for part in raw.split(";") if part.strip()]
    literals = []
    for part in parts:
        word_count = len(re.findall(r"[A-Za-z0-9']+", part))
        if word_count == 0 or word_count > 8 or INSTRUCTIONAL_EXCLUSION.search(part):
            return [], raw
        literals.append(normalize_phrase(part))
    return list(dict.fromkeys(literals)), None


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    baseline_dictionary = json.loads(DICTIONARY.read_text(encoding="utf-8"))

    icon_sheet = workbook["Icon_Mapping"]
    icon_rows = icon_sheet.iter_rows(values_only=True)
    icon_headers = next(icon_rows)
    icon_ix = {name: index for index, name in enumerate(icon_headers) if name}
    items_by_name = {}
    items_by_row = {}
    for row_number, row in enumerate(icon_rows, start=2):
        name = str(row[icon_ix["Icon_name"]]).strip()
        item_id = f"memory.{row_number:04d}_{slugify(name)}"
        keywords = [
            keyword.strip()
            for keyword in str(row[icon_ix["keywords_mapping"]] or "").split(";")
            if keyword.strip()
        ]
        item = {
            "id": item_id,
            "row_number": row_number,
            "displayName": name,
            "keywords": keywords,
            "visualConcept": str(row[icon_ix["visual_concept"]] or "").strip(),
            "category": str(row[icon_ix["Primary_Domain"]] or "Uncategorized").strip(),
            "tier": str(row[icon_ix["Tier"]] or "Tier 2").strip(),
        }
        items_by_name[name] = item
        items_by_row[row_number] = item
    if len(items_by_name) != ICON_COUNT:
        raise ValueError(f"Expected {ICON_COUNT} icons, found {len(items_by_name)}")

    reflect_sheet = workbook["Reflect_Icon_Map"]
    reflect_rows = reflect_sheet.iter_rows(values_only=True)
    reflect_headers = next(reflect_rows)
    reflect_ix = {name: index for index, name in enumerate(reflect_headers)}
    reflect_by_name = {}
    category_items = defaultdict(list)
    for row in reflect_rows:
        name = str(row[reflect_ix["Icon_name"]]).strip()
        category = str(row[reflect_ix["Primary_Category"]]).strip()
        display_order = int(row[reflect_ix["Display_Order"]])
        frequency = str(row[reflect_ix["Frequency_Level"]]).strip()
        reflect_by_name[name] = {"category": category, "order": display_order, "frequency": frequency}
        category_items[category].append((display_order, items_by_name[name]["id"]))
    if len(reflect_by_name) != ICON_COUNT:
        raise ValueError("Reflect_Icon_Map does not cover all icons")

    rarity_for_frequency = {"High": "common", "Medium": "common", "Low": "uncommon", "Rare": "rare"}
    runtime_items = {}
    for name, item in items_by_name.items():
        reflect = reflect_by_name[name]
        baseline = baseline_dictionary["items"].get(item["id"]) if baseline_dictionary else None
        if baseline_dictionary and (
            not baseline or baseline.get("displayName") != name
        ):
            raise ValueError(f"Stable item identity changed at {item['id']}: {name!r}")
        runtime_items[item["id"]] = {
            "displayName": name,
            "rarity": rarity_for_frequency[reflect["frequency"]],
            "category": reflect["category"],
            "bagsCategory": item["category"],
            "keywords": item["keywords"],
            "visualConcept": item["visualConcept"],
        }

    safety_sheet = workbook["Keyword_Safety"]
    safety_rows = safety_sheet.iter_rows(values_only=True)
    safety_headers = next(safety_rows)
    safety_ix = {name: index for index, name in enumerate(safety_headers) if name}
    candidates = defaultdict(list)
    never_auto = 0
    never_auto_keywords = set()
    unresolved_exclusions = []
    executable_rows = 0
    for excel_row, row in enumerate(safety_rows, start=2):
        # The workbook keeps formatting below the populated table, which openpyxl
        # exposes as fully empty rows. Ignore only rows with no rule content;
        # partially populated rows must still fail validation below.
        if not any(value is not None and str(value).strip() for value in row):
            continue
        source_name = str(row[safety_ix["Icon_name"]]).strip()
        name = ORPHAN_ICON_REMAP.get(source_name, source_name)
        if name not in items_by_name:
            raise ValueError(f"Keyword_Safety row {excel_row} references missing icon {source_name!r}")
        keyword = normalize_phrase(row[safety_ix["Keyword"]])
        mode = str(row[safety_ix["Trigger_Mode"]]).strip()
        if not keyword:
            continue
        if mode == "NEVER_AUTO":
            never_auto += 1
            never_auto_keywords.add(keyword)
            continue
        if mode not in {"AUTO", "AUTO_UNLESS_EXCLUDED"}:
            raise ValueError(f"Unknown Trigger_Mode at Keyword_Safety row {excel_row}: {mode}")
        exclusions = []
        if mode == "AUTO_UNLESS_EXCLUDED":
            exclusions, unresolved = parse_literal_exclusions(row[safety_ix["Exclusion_Rule"]])
            if unresolved:
                unresolved_exclusions.append({"row": excel_row, "icon": name, "keyword": keyword, "rule": unresolved})
        candidates[keyword].append({
            "name": name,
            "mode": mode,
            "exclusions": exclusions,
            "unsafe_instructional_exclusion": unresolved is not None if mode == "AUTO_UNLESS_EXCLUDED" else False,
        })
        executable_rows += 1

    cross_icon_conflicts = {
        keyword: sorted({candidate["name"] for candidate in options})
        for keyword, options in candidates.items()
        if len({candidate["name"] for candidate in options}) > 1
    }
    if set(cross_icon_conflicts) != set(CONFLICT_WINNERS):
        missing = sorted(set(cross_icon_conflicts) - set(CONFLICT_WINNERS))
        stale = sorted(set(CONFLICT_WINNERS) - set(cross_icon_conflicts))
        raise ValueError(f"Conflict winner table mismatch: missing={missing}, stale={stale}")

    synonyms = {}
    exclusions = {}
    conflict_report = []
    unsafe_keywords_omitted = []
    for keyword, options in sorted(candidates.items()):
        names = {option["name"] for option in options}
        winner_name = CONFLICT_WINNERS.get(keyword) if len(names) > 1 else next(iter(names))
        if winner_name not in names:
            raise ValueError(f"Invalid conflict winner for {keyword}: {winner_name} not in {sorted(names)}")
        winner_options = [
            option for option in options
            if option["name"] == winner_name and not option["unsafe_instructional_exclusion"]
        ]
        if len(names) > 1:
            conflict_report.append({
                "keyword": keyword,
                "candidates": sorted(names),
                "winner": winner_name,
                "auto_active": bool(winner_options),
            })
        if not winner_options:
            unsafe_keywords_omitted.append(keyword)
            continue
        synonyms[keyword] = items_by_name[winner_name]["id"]
        literal_exclusions = sorted({rule for option in winner_options for rule in option["exclusions"]})
        if literal_exclusions:
            exclusions[keyword] = literal_exclusions

    dictionary = {"items": runtime_items, "synonyms": synonyms, "exclusions": exclusions}
    DICTIONARY.write_text(json.dumps(dictionary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    image_lines = [
        "/**",
        " * GENERATED by tools/build-item-data-v19.py — DO NOT EDIT.",
        f" * One bundled 256px WebP per v31 memory item ({ICON_COUNT:,} total).",
        " */",
        "",
        "export const ITEM_IMAGES: Record<string, number> = {",
    ]
    for row_number in sorted(items_by_row):
        item = items_by_row[row_number]
        image_lines.append(
            f"  {json.dumps(item['id'])}: require('../../assets/items/each/{item['id']}.webp'),"
        )
    image_lines.extend(["};", ""])
    if WRITE_IMAGE_MAP:
        IMAGE_MAP.write_text("\n".join(image_lines), encoding="utf-8")

    category_sheet = workbook["Reflect_Categories"]
    category_rows = category_sheet.iter_rows(values_only=True)
    category_headers = next(category_rows)
    category_ix = {name: index for index, name in enumerate(category_headers)}
    categories = []
    for row in category_rows:
        label = str(row[category_ix["Reflect_Category"]]).strip()
        if label not in CATEGORY_KEYS:
            raise ValueError(f"Missing stable category key for {label}")
        ordered_ids = [item_id for _, item_id in sorted(category_items[label])]
        categories.append({
            "order": int(row[category_ix["Category_Order"]]),
            "key": CATEGORY_KEYS[label],
            "label": label,
            "question": str(row[category_ix["Reflect_Question"]]).strip(),
            "itemIds": ordered_ids,
        })
    categories.sort(key=lambda category: category["order"])
    if len(categories) != 12 or sum(len(category["itemIds"]) for category in categories) != ICON_COUNT:
        raise ValueError("Reflect category generation is incomplete")

    # v31 Guided Prompt secondary tabs. First appearance determines tab order;
    # row order determines icon order. Categories intentionally absent from the
    # sheet render their original full grid with no empty tab bar.
    subcategory_sheet = workbook["Reflect_Subcategory_Map"]
    subcategory_rows = subcategory_sheet.iter_rows(values_only=True)
    subcategory_headers = next(subcategory_rows)
    subcategory_ix = {name: index for index, name in enumerate(subcategory_headers) if name}
    subcategories = defaultdict(dict)
    subcategory_seen_icons = set()
    for excel_row, row in enumerate(subcategory_rows, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        main_label = str(row[subcategory_ix["Main_Category"]] or "").strip()
        secondary_label = str(row[subcategory_ix["Secondary_Category"]] or "").strip()
        name = str(row[subcategory_ix["Icon_name"]] or "").strip()
        if main_label not in CATEGORY_KEYS or not secondary_label or name not in items_by_name:
            raise ValueError(f"Invalid Reflect_Subcategory_Map row {excel_row}")
        if name in subcategory_seen_icons:
            raise ValueError(f"Duplicate subcategory icon at row {excel_row}: {name}")
        if reflect_by_name[name]["category"] != main_label:
            raise ValueError(
                f"Subcategory row {excel_row} puts {name!r} in {main_label!r}, "
                f"but Reflect_Icon_Map uses {reflect_by_name[name]['category']!r}"
            )
        subcategory_seen_icons.add(name)
        subcategories[main_label].setdefault(secondary_label, []).append(items_by_name[name]["id"])

    for category in categories:
        category["subcategories"] = [
            {"key": slugify(label), "label": label, "itemIds": item_ids}
            for label, item_ids in subcategories.get(category["label"], {}).items()
        ]
        mapped = sum(len(subcategory["itemIds"]) for subcategory in category["subcategories"])
        if category["subcategories"] and mapped != len(category["itemIds"]):
            raise ValueError(
                f"Subcategory coverage mismatch for {category['label']}: {mapped} vs {len(category['itemIds'])}"
            )

    guided_lines = [
        "/** GENERATED by tools/build-item-data-v19.py — DO NOT EDIT. */",
        "export interface PromptSubcategoryDef { key: string; label: string; itemIds: string[] }",
        "export interface PromptCategoryDef { key: string; label: string; question: string; itemIds: string[]; subcategories: PromptSubcategoryDef[] }",
        "",
        "export const PROMPT_CATEGORIES: PromptCategoryDef[] = [",
    ]
    for category in categories:
        # Subcategorized groups already contain every id in workbook order.
        # Avoid bundling the same ~5k strings twice; guided-prompts.ts derives
        # the complete first-level list by flattening these groups.
        bundled_item_ids = [] if category["subcategories"] else category["itemIds"]
        guided_lines.append(
            "  { key: %s, label: %s, question: %s, itemIds: ["
            % (json.dumps(category["key"]), json.dumps(category["label"]), json.dumps(category["question"]))
        )
        guided_lines.extend(f"    {json.dumps(item_id)}," for item_id in bundled_item_ids)
        guided_lines.append("  ], subcategories: [")
        for subcategory in category["subcategories"]:
            guided_lines.append(
                "    { key: %s, label: %s, itemIds: ["
                % (json.dumps(subcategory["key"]), json.dumps(subcategory["label"]))
            )
            guided_lines.extend(f"      {json.dumps(item_id)}," for item_id in subcategory["itemIds"])
            guided_lines.append("    ] },")
        guided_lines.append("  ] },")
    guided_lines.append("];")
    GUIDED_CATALOG.write_text("\n".join(guided_lines) + "\n", encoding="utf-8")

    qa = {
        "schema": "memory-items-v32-data-qa@1",
        "source_workbook": WORKBOOK.name,
        "items": len(runtime_items),
        "executable_keyword_rows": executable_rows,
        "unique_executable_keywords": len(synonyms),
        "never_auto_rows_omitted": never_auto,
        "never_auto_unique_keywords": len(never_auto_keywords),
        "never_auto_keywords_with_explicit_auto_carrier": sorted(never_auto_keywords & set(candidates)),
        "auto_unless_excluded_keywords_with_literal_rules": len(exclusions),
        "unresolved_instructional_exclusions": len(unresolved_exclusions),
        "unsafe_auto_unless_excluded_keywords_omitted": len(unsafe_keywords_omitted),
        "unsafe_keyword_samples": unsafe_keywords_omitted[:100],
        "unresolved_instructional_exclusion_samples": unresolved_exclusions[:100],
        "cross_icon_conflicts": conflict_report,
        "reflect_categories": [
            {
                "order": category["order"], "key": category["key"], "label": category["label"],
                "items": len(category["itemIds"]),
                "subcategories": [
                    {"key": subcategory["key"], "label": subcategory["label"], "items": len(subcategory["itemIds"])}
                    for subcategory in category.get("subcategories", [])
                ],
            }
            for category in categories
        ],
    }
    QA_PATH.write_text(json.dumps(qa, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {len(runtime_items)} items, {len(synonyms)} executable keywords, {len(categories)} categories")
    print(f"resolved {len(conflict_report)} cross-icon keyword conflicts")
    print(f"literal exclusion maps: {len(exclusions)}; instructional exclusions flagged: {len(unresolved_exclusions)}")


if __name__ == "__main__":
    main()

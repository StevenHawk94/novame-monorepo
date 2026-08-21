#!/usr/bin/env python3
"""Items v3 pipeline (2026-07-30): icon detection, slicing, and data gen.

Input:  tools/item-source/memory-items/*.png (42 pages, near-white bg,
        icons in reading order but NOT on a uniform grid) +
        icon_keyword_mapping_final.xlsx (master: 1693 items with sequence
        `n`, bags category, keywords; sheet 2: 11 prompt categories).

Detection instead of grid math: threshold the near-white background, dilate
to glue icon fragments, label connected components (scipy), cluster into
row bands by y-center, sort bands top-to-bottom / icons left-to-right, and
concatenate pages in file order. That reading order must equal the master's
`n` order per category — a count mismatch fails the page loudly rather than
slicing garbage.

Background removal that cannot eat icon content: flood fill the near-white
region FROM THE CROP BORDER only — exterior becomes transparent, interior
whites (paper, teeth, highlights) stay. The crop is the component bbox plus
6% padding, clamped to the page.

Outputs:
  apps/mobile/assets/items/each/<itemId>.webp        1693 icons, 256px
  packages/engine/src/items/dictionary.json          items + synonyms
  apps/mobile/src/lib/item-images.g.ts               require map
  apps/mobile/src/lib/guided-catalog.g.ts            11 prompt categories
  supabase/migrations/<latest>_items_catalog.sql    catalog upsert
  <scratch>/preview.html + flagged list              human spot-check
"""
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl
from PIL import Image
from scipy import ndimage

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / 'tools' / 'item-source' / 'memory-items'
OUT_ICONS = ROOT / 'apps' / 'mobile' / 'assets' / 'items' / 'each'
SCRATCH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('/tmp/items-v3')
SCRATCH.mkdir(parents=True, exist_ok=True)

CANVAS = 256
SAFE = 0.10          # margin inside the canvas
PAD = 0.06           # padding around the detected bbox before cropping
BG_THRESH = 26       # color distance from page background to count as content
MIN_AREA_FRAC = 0.06 # of median component area — below = noise, merge/drop

def slugify(s: str) -> str:
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    s = re.sub(r"[^a-z0-9]+", '_', s.lower()).strip('_')
    return s

# ---------------- master data ----------------
wb = openpyxl.load_workbook(SRC / 'icon_keyword_mapping_updated.xlsx', read_only=True)
ws = wb['icon_keyword_mapping_final']
master = []  # per category, ordered by n
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[5]:
        continue
    master.append({
        'bags': str(r[0]).strip(),
        'category': str(r[1]).strip(),
        'n': int(r[2]),
        'name': str(r[5]).strip(),
        'keywords': [k.strip().lower() for k in re.split(r'[;,]', str(r[6])) if k.strip()],
    })
by_cat = defaultdict(list)
for m in master:
    by_cat[m['category']].append(m)
for cat in by_cat:
    by_cat[cat].sort(key=lambda m: m['n'])
    for m in by_cat[cat]:
        m['id'] = f"{slugify(m['category'])}.{slugify(m['name'])}"

ids = [m['id'] for m in master]
assert len(ids) == len(set(ids)), 'duplicate item ids after slugify'
print(f'master: {len(master)} items, {len(by_cat)} categories')

# ---------------- detection ----------------
def detect_icons(im: Image.Image, iters: int):
    """Return list of (x0, y0, x1, y1) bboxes in reading order."""
    rgb = np.asarray(im.convert('RGB'), dtype=np.int16)
    h, w, _ = rgb.shape
    # page background = median of the border ring
    ring = np.concatenate([
        rgb[0:6].reshape(-1, 3), rgb[-6:].reshape(-1, 3),
        rgb[:, 0:6].reshape(-1, 3), rgb[:, -6:].reshape(-1, 3),
    ])
    bg = np.median(ring, axis=0)
    dist = np.abs(rgb - bg).sum(axis=2)
    mask = dist > BG_THRESH
    # glue icon fragments (dangling key rings, steam wisps) before labeling
    glued = ndimage.binary_dilation(mask, iterations=iters)
    labels, count = ndimage.label(glued)
    if count == 0:
        return []
    objs = ndimage.find_objects(labels)
    comps = []
    areas = []
    for sl in objs:
        sub = mask[sl]
        area = int(sub.sum())
        if area == 0:
            continue
        comps.append((sl[1].start, sl[0].start, sl[1].stop, sl[0].stop, area))
        areas.append(area)
    med = float(np.median(areas))
    comps = [c for c in comps if c[4] >= med * MIN_AREA_FRAC]
    # cluster into row bands by y-center with a gap threshold
    comps.sort(key=lambda c: (c[1] + c[3]) / 2)
    bands = []
    for c in comps:
        yc = (c[1] + c[3]) / 2
        if bands and yc - bands[-1]['yc'] < (c[3] - c[1]) * 0.6:
            b = bands[-1]
            b['items'].append(c)
            b['yc'] = sum((i[1] + i[3]) / 2 for i in b['items']) / len(b['items'])
        else:
            bands.append({'yc': yc, 'items': [c]})
    out = []
    for b in bands:
        b['items'].sort(key=lambda c: c[0])
        out.extend((c[0], c[1], c[2], c[3]) for c in b['items'])
    return out

def cut_icon(im: Image.Image, bbox, out_path: Path):
    """Crop with padding, flood-fill exterior bg to transparent, center on canvas."""
    W, H = im.size
    x0, y0, x1, y1 = bbox
    pw, ph = int((x1 - x0) * PAD) + 2, int((y1 - y0) * PAD) + 2
    x0, y0 = max(0, x0 - pw), max(0, y0 - ph)
    x1, y1 = min(W, x1 + pw), min(H, y1 + ph)
    crop = im.convert('RGB').crop((x0, y0, x1, y1))
    arr = np.asarray(crop, dtype=np.int16)
    ring = np.concatenate([
        arr[0:3].reshape(-1, 3), arr[-3:].reshape(-1, 3),
        arr[:, 0:3].reshape(-1, 3), arr[:, -3:].reshape(-1, 3),
    ])
    bg = np.median(ring, axis=0)
    nearbg = (np.abs(arr - bg).sum(axis=2) <= BG_THRESH)
    # exterior = near-bg pixels connected to the border (interior whites survive)
    lab, _ = ndimage.label(nearbg)
    border_labels = set(np.unique(np.concatenate([lab[0], lab[-1], lab[:, 0], lab[:, -1]])))
    border_labels.discard(0)
    exterior = np.isin(lab, list(border_labels)) if border_labels else np.zeros_like(nearbg)
    alpha = np.where(exterior, 0, 255).astype(np.uint8)
    # Drop slivers of neighboring icons that leak across grid-cell borders:
    # keep only content components with meaningful area (or near the center),
    # then bbox the survivors. Detached icon parts (steam, sparks) survive the
    # 6% threshold; a neighbor's clipped edge doesn't get to widen the frame.
    content = alpha > 0
    clab, ccount = ndimage.label(content)
    if ccount > 1:
        sizes = ndimage.sum(content, clab, range(1, ccount + 1))
        biggest = float(sizes.max())
        keep = {i + 1 for i, sz in enumerate(sizes) if sz >= biggest * 0.06}
        drop_mask = ~np.isin(clab, list(keep)) & content
        alpha[drop_mask] = 0
    rgba = np.dstack([np.asarray(crop, dtype=np.uint8), alpha])
    icon = Image.fromarray(rgba, 'RGBA')
    # tight bbox on alpha, then center on the canvas with the safe margin
    abox = icon.getchannel('A').getbbox()
    if abox:
        icon = icon.crop(abox)
    inner = int(CANVAS * (1 - 2 * SAFE))
    iw, ih = icon.size
    scale = min(inner / iw, inner / ih)
    icon = icon.resize((max(1, int(iw * scale)), max(1, int(ih * scale))), Image.LANCZOS)
    canvas = Image.new('RGBA', (CANVAS, CANVAS), (0, 0, 0, 0))
    canvas.paste(icon, ((CANVAS - icon.size[0]) // 2, (CANVAS - icon.size[1]) // 2), icon)
    canvas.save(out_path, 'WEBP', quality=92, method=4)

FUSED_LOG: list = []


def banded_grid_detect(im: Image.Image, cols: int = 8, lenient_dedupe: bool = False):
    """Fallback for dense uniform-grid pages where neighboring icons touch:
    row bands come from component clustering (robust even when icons within a
    row merge), then each band is split into `cols` uniform columns and empty
    cells are skipped. Reading order is band-major."""
    rgb = np.asarray(im.convert('RGB'), dtype=np.int16)
    h, w, _ = rgb.shape
    ring = np.concatenate([
        rgb[0:6].reshape(-1, 3), rgb[-6:].reshape(-1, 3),
        rgb[:, 0:6].reshape(-1, 3), rgb[:, -6:].reshape(-1, 3),
    ])
    bg = np.median(ring, axis=0)
    mask = np.abs(rgb - bg).sum(axis=2) > BG_THRESH
    # row bands from the horizontal projection: rows of content separated by
    # (near-)empty gaps
    proj = mask.sum(axis=1)
    on = proj > (w * 0.01)
    bands = []
    start = None
    for y, v in enumerate(on):
        if v and start is None:
            start = y
        elif not v and start is not None:
            if y - start > 12:
                bands.append((start, y))
            start = None
    if start is not None and h - start > 12:
        bands.append((start, h))
    # A hairline gap inside one icon (a cord under a power strip) can strand
    # a fragment as its own tiny band. Absorb ONLY runt bands (well under the
    # median height) into their nearest neighbor — full-height rows stay
    # separate no matter how thin the gap (dense pages have 4-6px row gaps).
    if len(bands) > 1:
        med_h = sorted(b[1] - b[0] for b in bands)[len(bands) // 2]
        absorbed = []
        for b in bands:
            if absorbed and (b[1] - b[0]) < med_h * 0.4:
                absorbed[-1] = (absorbed[-1][0], b[1])
            elif absorbed and (absorbed[-1][1] - absorbed[-1][0]) < med_h * 0.4:
                absorbed[-1] = (absorbed[-1][0], b[1])
            else:
                absorbed.append(b)
        bands = absorbed
    # Rows that touch vertically fuse into one double-height band — split any
    # band ~k× the median height into k equal sub-bands.
    if bands:
        heights = sorted(b[1] - b[0] for b in bands)
        med = heights[len(heights) // 2]
        split = []
        for (y0, y1) in bands:
            k = max(1, round((y1 - y0) / med)) if med > 0 else 1
            if k <= 1:
                split.append((y0, y1))
            else:
                step = (y1 - y0) / k
                for i in range(k):
                    split.append((int(y0 + i * step), int(y0 + (i + 1) * step)))
        bands = split
    cw = w / cols
    out = []
    for (y0, y1) in bands:
        band_mask = mask[y0:y1]
        blab, _ = ndimage.label(band_mask)
        cell_area = (y1 - y0) * cw
        min_px = max(60, int(cell_area * 0.015))
        row_cells = []  # [col, bbox, dom_label, dom_frac]
        for cidx in range(cols):
            x0, x1 = int(cidx * cw), int((cidx + 1) * cw)
            cell = band_mask[:, x0:x1]
            px = int(cell.sum())
            if px < min_px:
                continue
            labs = blab[:, x0:x1][cell]
            dom = int(np.bincount(labs).argmax()) if labs.size else 0
            dom_frac = float((labs == dom).sum()) / max(1, labs.size)
            sel = (blab[:, x0:x1] == dom) & cell
            ys, xs = np.nonzero(sel)
            if xs.size == 0:
                ys, xs = np.nonzero(cell)
            bbox = (x0 + int(xs.min()), y0 + int(ys.min()),
                    x0 + int(xs.max()) + 1, y0 + int(ys.max()) + 1)
            row_cells.append([cidx, bbox, dom, dom_frac])
        def seam_connected(border_x: float) -> bool:
            sx0 = max(0, int(border_x - cw * 0.18))
            sx1 = min(band_mask.shape[1], int(border_x + cw * 0.18))
            win = band_mask[:, sx0:sx1]
            colsum = win.sum(axis=0)
            if colsum.size == 0:
                return False
            return int(colsum.min()) > max(2, int(colsum.max() * 0.05))
        # group adjacent cells sharing a dominant component
        groups = []
        for rc in row_cells:
            if (groups and rc[2] == groups[-1][-1][2] and rc[2] != 0
                    and rc[0] == groups[-1][-1][0] + 1):
                groups[-1].append(rc)
            else:
                groups.append([rc])
        for g in groups:
            if len(g) == 1:
                out.append(tuple(g[0][1]))
                continue
            if (lenient_dedupe and all(rc[3] >= 0.6 for rc in g)
                    and all(seam_connected(g[i + 1][0] * cw) for i in range(len(g) - 1))):
                ys, xs = np.nonzero(blab == g[0][2])
                out.append((int(xs.min()), y0 + int(ys.min()),
                            int(xs.max()) + 1, y0 + int(ys.max()) + 1))
                continue
            # One component dominates several windows (a drifted icon straddling
            # the border can drown its neighbor). Re-assign by the REAL
            # components inside the group's x-range: majors sorted by x-center,
            # one per cell when the counts agree.
            FUSED_LOG.append((getattr(im, 'filename', '?'), y0, g[0][0]))
            gx0 = int(g[0][0] * cw)
            gx1 = min(band_mask.shape[1], int((g[-1][0] + 1) * cw))
            in_range = {}
            for lbl in np.unique(blab[:, gx0:gx1]):
                if lbl == 0:
                    continue
                ys, xs = np.nonzero(blab == lbl)
                inside = ((xs >= gx0) & (xs < gx1)).sum()
                if inside > 0:
                    in_range[int(lbl)] = (int(inside), float(xs.mean()))
            biggest = max(v[0] for v in in_range.values())
            majors = sorted(
                [(v[1], lbl) for lbl, v in in_range.items() if v[0] >= biggest * 0.2])
            if len(majors) == len(g):
                for _xc, lbl in majors:
                    ys, xs = np.nonzero(blab == lbl)
                    out.append((int(xs.min()), y0 + int(ys.min()),
                                int(xs.max()) + 1, y0 + int(ys.max()) + 1))
                continue
            # fallback: cut the shared component at its thinnest valley
            # between cell centers
            comp_cols = (blab == g[0][2]).sum(axis=0)
            cuts = []
            for i in range(len(g) - 1):
                c0 = int((g[i][0] + 0.5) * cw)
                c1 = int((g[i + 1][0] + 0.5) * cw)
                lo, hi = max(0, c0), min(len(comp_cols), c1)
                seg = comp_cols[lo:hi]
                cuts.append(lo + int(np.argmin(seg)) if seg.size else int(g[i + 1][0] * cw))
            edges = [0] + cuts + [band_mask.shape[1]]
            for i in range(len(g)):
                win = blab[:, edges[i]:edges[i + 1]] == g[i][2]
                ys, xs = np.nonzero(win)
                if xs.size == 0:
                    out.append(tuple(g[i][1]))
                    continue
                out.append((edges[i] + int(xs.min()), y0 + int(ys.min()),
                            edges[i] + int(xs.max()) + 1, y0 + int(ys.max()) + 1))
    return out


def merge_nearest(boxes, target):
    """Deterministic fallback: fuse the smallest box into its nearest
    neighbor until len(boxes) == target (over-detection only)."""
    boxes = [list(b) for b in boxes]
    def area(b): return (b[2] - b[0]) * (b[3] - b[1])
    def center(b): return ((b[0] + b[2]) / 2, (b[1] + b[3]) / 2)
    while len(boxes) > target:
        i = min(range(len(boxes)), key=lambda k: area(boxes[k]))
        ci = center(boxes[i])
        j = min((k for k in range(len(boxes)) if k != i),
                key=lambda k: (center(boxes[k])[0] - ci[0]) ** 2 + (center(boxes[k])[1] - ci[1]) ** 2)
        a, b = boxes[i], boxes[j]
        b[0], b[1] = min(a[0], b[0]), min(a[1], b[1])
        b[2], b[3] = max(a[2], b[2]), max(a[3], b[3])
        boxes.pop(i)
    # restore reading order: band by y-center then x
    boxes.sort(key=lambda b: (b[1] + b[3]) / 2)
    bands = []
    for b in boxes:
        yc = (b[1] + b[3]) / 2
        if bands and yc - bands[-1]['yc'] < (b[3] - b[1]) * 0.6:
            band = bands[-1]
            band['items'].append(b)
            band['yc'] = sum((x[1] + x[3]) / 2 for x in band['items']) / len(band['items'])
        else:
            bands.append({'yc': yc, 'items': [b]})
    out = []
    for band in bands:
        band['items'].sort(key=lambda b: b[0])
        out.extend(tuple(b) for b in band['items'])
    return out


def detect_faces(im: Image.Image, expected: int):
    """Emotions page: big round faces with caption text beneath and a page
    heading. Keep only large, roughly-square components (the faces); text
    lines are wide/short or tiny and get dropped, so captions never end up
    inside the sliced icon."""
    boxes = detect_icons(im, 5)
    if not boxes:
        return []
    amax = max((b[2] - b[0]) * (b[3] - b[1]) for b in boxes)
    faces = []
    for b in boxes:
        w, h = b[2] - b[0], b[3] - b[1]
        ratio = w / h if h else 99
        if 0.7 <= ratio <= 1.45 and w * h >= amax * 0.4:
            faces.append(b)
    return faces


SPECIAL_DETECT = {'Emotions & Expressions': detect_faces}


def pages_for(cat: str):
    pages = sorted(SRC.glob(f'{cat}-*.png'),
                   key=lambda p: int(p.stem.rsplit('-', 1)[1]))
    return pages or [SRC / f'{cat}.png']

if __name__ == '__main__':
    OUT_ICONS.mkdir(parents=True, exist_ok=True)
    # wipe stale icons so the set exactly mirrors the master
    for old in OUT_ICONS.glob('*.webp'):
        old.unlink()

    failures = []
    done = 0
    # Down to 1: dense pages (Food & Drink v2, 2026-08-05) already glue
    # neighboring icons at 5, so the ladder must be able to split first.
    ITER_LADDER = [1, 2, 3, 5, 8, 11, 15, 20, 26, 33]
    for cat, items in by_cat.items():
        ims = [Image.open(p) for p in pages_for(cat)]
        if cat in SPECIAL_DETECT:
            boxes_all = [(im, SPECIAL_DETECT[cat](im, len(items))) for im in ims]
            detected = sum(len(b) for _, b in boxes_all)
            if detected != len(items):
                failures.append((cat, len(items), detected))
                print(f'  MISMATCH {cat} (special): expected {len(items)}, got {detected}')
                continue
            i = 0
            for im, boxes in boxes_all:
                for bbox in boxes:
                    cut_icon(im, bbox, OUT_ICONS / f"{items[i]['id']}.webp")
                    i += 1
                    done += 1
            print(f'  {cat}: {len(items)} ok (special face detect)')
            continue
        chosen = None
        best_over = None  # (overshoot, boxes_all) — smallest overshoot wins
        # adaptive glue: climb the dilation ladder until the category's icon
        # count matches the master exactly; keep the tightest over-detection
        # as merge fodder if no rung is exact
        for iters in ITER_LADDER:
            boxes_all = [(im, detect_icons(im, iters)) for im in ims]
            detected = sum(len(b) for _, b in boxes_all)
            if detected == len(items):
                chosen = (boxes_all, iters, 'exact')
                break
            if detected > len(items):
                over = detected - len(items)
                if best_over is None or over < best_over[0]:
                    best_over = (over, boxes_all)
        if chosen is None:
            # Component counting couldn't hit the exact total — these are the
            # dense uniform-grid pages where neighbors touch. Slice by row
            # bands × uniform columns instead (alignment-safe by design).
            # Celebration's bunting garland genuinely spans two slots — that
            # page gets the lenient wide-icon dedupe; dense catalogs (packed
            # food rows) need the strict seam test.
            lenient = cat == 'Celebration & Gifts'
            boxes_all = [(im, banded_grid_detect(im, lenient_dedupe=lenient)) for im in ims]
            detected = sum(len(b) for _, b in boxes_all)
            if detected != len(items):
                failures.append((cat, len(items), f'grid fallback got {detected}'))
                print(f'  MISMATCH {cat}: grid fallback expected {len(items)}, got {detected}')
                continue
            chosen = (boxes_all, 0, 'banded-grid')
        boxes_all, iters, how = chosen
        detected = sum(len(b) for _, b in boxes_all)
        if detected != len(items):
            failures.append((cat, len(items), detected))
            print(f'  MISMATCH {cat}: expected {len(items)}, got {detected}')
            continue
        i = 0
        for im, boxes in boxes_all:
            for bbox in boxes:
                cut_icon(im, bbox, OUT_ICONS / f"{items[i]['id']}.webp")
                i += 1
                done += 1
        print(f'  {cat}: {len(items)} ok (iters={iters}, {how})')

    print(f'\nsliced {done}/{len(master)}; category failures: {len(failures)}')
    for f in failures:
        print('  FAILED:', f)
    if failures:
        sys.exit(1)

    # ---------------- generated artifacts ----------------
    # 1) engine dictionary (items + synonyms). Keyword collisions: first n wins.
    items_out, synonyms, collisions = {}, {}, []
    for m in master:
        items_out[m['id']] = {
            'category': slugify(m['category']),
            'displayName': m['name'],
            'rarity': 'common',
            'bagsCategory': m['bags'],
        }
        for kw in m['keywords']:
            if kw in synonyms and synonyms[kw] != m['id']:
                collisions.append((kw, synonyms[kw], m['id']))
                continue
            synonyms[kw] = m['id']
    dict_path = ROOT / 'packages' / 'engine' / 'src' / 'items' / 'dictionary.json'
    dict_path.write_text(json.dumps({'items': items_out, 'synonyms': synonyms},
                                    indent=1, ensure_ascii=False, sort_keys=True))
    print(f'dictionary: {len(items_out)} items, {len(synonyms)} synonyms, '
          f'{len(collisions)} keyword collisions (first n kept)')
    (SCRATCH / 'keyword-collisions.txt').write_text(
        '\n'.join(f'{k}: kept {a}, dropped {b}' for k, a, b in collisions))

    # 2) mobile image map
    lines = [
        '/**',
        ' * GENERATED by tools/build-items-v3.py — DO NOT EDIT.',
        f' * One 256px webp per memory item ({len(master)} items, v3 catalog).',
        ' * Re-run the script after icon_keyword_mapping_final.xlsx changes.',
        ' */', '',
        'export const ITEM_IMAGES: Record<string, number> = {',
    ]
    for m in sorted(master, key=lambda m: m['id']):
        lines.append(f"  '{m['id']}': require('../../assets/items/each/{m['id']}.webp'),")
    lines += ['};', '']
    (ROOT / 'apps' / 'mobile' / 'src' / 'lib' / 'item-images.g.ts').write_text('\n'.join(lines))

    # 3) guided catalog: 11 prompt categories with ranked item ids
    ws2 = wb['prompt reflection']
    by_name = {m['name'].strip().lower(): m['id'] for m in master}
    prompt_cats = {}
    for r in ws2.iter_rows(min_row=2, values_only=True):
        if not r[2]:
            continue
        cat2 = str(r[0]).strip()
        iid = by_name.get(str(r[2]).strip().lower())
        if not iid:
            continue
        prompt_cats.setdefault(cat2, []).append((int(r[1] or 0), iid))
    # Food & Drink shows the FULL master category, not the curated subset
    # (user call 2026-07-31) — master n order.
    prompt_cats['Food & Drink'] = [(m['n'], m['id']) for m in by_cat['Food & Drink']]
    glines = [
        '/**',
        ' * GENERATED by tools/build-items-v3.py — DO NOT EDIT.',
        ' * The 11 prompt-reflection categories (curated, ranked subsets of the',
        ' * master catalog) used by Guided Prompts and Object Reflect.',
        ' */', '',
        'export interface PromptCategoryDef { key: string; label: string; itemIds: string[] }', '',
        'export const PROMPT_CATEGORIES: PromptCategoryDef[] = [',
    ]
    for cat2, pairs in prompt_cats.items():
        pairs.sort()
        glines.append(f"  {{ key: '{slugify(cat2)}', label: {json.dumps(cat2)}, itemIds: [")
        for _, iid in pairs:
            glines.append(f"    '{iid}',")
        glines.append('  ] },')
    glines += ['];', '']
    (ROOT / 'apps' / 'mobile' / 'src' / 'lib' / 'guided-catalog.g.ts').write_text('\n'.join(glines))
    print(f'guided catalog: {len(prompt_cats)} categories, '
          f'{sum(len(v) for v in prompt_cats.values())} entries')

    # 4) SQL migration: upsert the v3 catalog (old rows stay for history)
    sql = [
        f'-- 031: items catalog refresh (2026-08-05). {len(master)} items from',
        '-- icon_keyword_mapping_updated.xlsx (Food & Drink expanded to 1208).',
        '-- Upsert only — earlier ids keep their rows.',
        '-- v3 renders per-item images; the v2 sprite-sheet coordinate columns',
        '-- become nullable legacy fields.',
        'ALTER TABLE public.items ALTER COLUMN sheet_id DROP NOT NULL;',
        'ALTER TABLE public.items ALTER COLUMN "row" DROP NOT NULL;',
        'ALTER TABLE public.items ALTER COLUMN col DROP NOT NULL;',
        '',
        'INSERT INTO public.items (id, display_name, category, rarity) VALUES',
    ]
    vals = []
    for m in sorted(master, key=lambda m: m['id']):
        name = m['name'].replace("'", "''")
        vals.append(f"  ('{m['id']}', '{name}', '{slugify(m['category'])}', 'common')")
    sql.append(',\n'.join(vals))
    sql.append('ON CONFLICT (id) DO UPDATE SET display_name = EXCLUDED.display_name,')
    sql.append('  category = EXCLUDED.category, rarity = EXCLUDED.rarity;')
    (ROOT / 'supabase' / 'migrations' / '20260805000031_items_food_expansion.sql').write_text('\n'.join(sql) + '\n')
    print('migration 031 written')

    # 5) preview page for the human spot-check
    rows_html = []
    for m in master:
        rows_html.append(
            f"<div class='c'><img src='../../../../../../Users/nihao/Desktop/Github/novame/apps/mobile/assets/items/each/{m['id']}.webp'>"
            f"<span>{m['name']}</span></div>")
    (SCRATCH / 'preview.html').write_text(
        "<meta charset='utf-8'><style>body{background:#f2e6cb;font-family:sans-serif}"
        ".c{display:inline-block;width:110px;text-align:center;margin:4px;background:#fff;"
        "border-radius:10px;padding:6px}img{width:84px;height:84px}span{font-size:10px;display:block}"
        "</style>" + '\n'.join(rows_html))
    print(f'preview: {SCRATCH / "preview.html"}')
    (SCRATCH / 'fused-cells.txt').write_text(
        '\n'.join(f'{f} band_y0={y} col={c}' for f, y, c in FUSED_LOG))
    print(f'fused-cell flags: {len(FUSED_LOG)} (see fused-cells.txt)')
